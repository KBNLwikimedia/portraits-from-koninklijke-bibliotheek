"""
Portraits from Koninklijke Bibliotheek — depicts Q5 (human) status analyser
============================================================================

Analyses structured data on Wikimedia Commons files in the category
"Portraits from Koninklijke Bibliotheek". For each file it checks
whether a depicts (P180) statement exists and, if so, whether the
depicted item is an instance of human (Q5). Non-human depicts values
(places, events, etc.) are discarded.

For every file the script also collects:
- English and Dutch file captions (MediaInfo labels on Commons).
- Non-hidden Commons categories.
- Global usage on other wikis (with Wikidata Q-id and label).

Output (one row per depicted human, so files with multiple depicts
values are split into multiple rows):
- kb-portraits-depicts-q5-status-all-YYYYMMDD.csv /
  kb-portraits-depicts-q5-status-missing-YYYYMMDD.csv — flat CSV files.
- kb-portraits-depicts-q5-status-YYYYMMDD.xlsx — Excel workbook with
  three sheets (all, missing, present).

The workflow:
1. Authenticate with Wikimedia Commons (credentials from .env).
2. Fetch all files from the category via the MediaWiki API.
3. Check P180 (depicts) and retrieve EN/NL file captions via
   wbgetentities on Commons (uses "statements" for claims and
   "labels" for captions on MediaInfo entities).
4. For every unique depicts Q-id, query Wikidata in a single pass to
   determine instance-of-human (P31=Q5) and retrieve English + Dutch
   labels and descriptions.
5. Filter depicts_map to humans only.
6. Fetch categories and global usage for all files.
7. Fetch English labels for Wikidata-usage Q-ids.
8. Build result rows (one per depicted human per file).
9. Write CSV and Excel output.

Requirements:
    - Python 3.7+
    - requests        (pip install requests)
    - python-dotenv   (pip install python-dotenv)
    - openpyxl        (pip install openpyxl)

Usage:
    python portraits-from-kb-assess-depicts-q5-status.py
"""

import os
import requests
import time
import csv
from datetime import date
from dotenv import load_dotenv
from openpyxl import Workbook

# Load credentials from .env file
load_dotenv()

COMMONS_API = "https://commons.wikimedia.org/w/api.php"
WIKIDATA_API = "https://www.wikidata.org/w/api.php"

# Get credentials from environment
USERNAME = os.getenv("COMMONS_USERNAME")
PASSWORD = os.getenv("COMMONS_PASSWORD")
USER_AGENT = os.getenv("COMMONS_USER_AGENT", "PortraitsFromKB_AssessmentBot/1.0")


def create_authenticated_session():
    """Create an authenticated MediaWiki session for Wikimedia Commons.

    Uses the COMMONS_USERNAME and COMMONS_PASSWORD environment variables
    to perform a bot-style login via the MediaWiki Action API.

    Returns:
        requests.Session on success, or None if login fails.
    """
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})

    # Step 1: Get login token
    params = {
        "action": "query",
        "meta": "tokens",
        "type": "login",
        "format": "json",
    }
    resp = session.get(COMMONS_API, params=params).json()
    login_token = resp["query"]["tokens"]["logintoken"]

    # Step 2: Log in
    data = {
        "action": "login",
        "lgname": USERNAME,
        "lgpassword": PASSWORD,
        "lgtoken": login_token,
        "format": "json",
    }
    resp = session.post(COMMONS_API, data=data).json()

    if resp.get("login", {}).get("result") == "Success":
        print(f"  Logged in as {resp['login']['lgusername']}")
        return session
    else:
        print(f"  Login failed: {resp}")
        return None


def get_category_members(session, category="Category:Portraits from Koninklijke Bibliotheek"):
    """Fetch all file members from a Wikimedia Commons category.

    Uses generator=categorymembers with gcmtype=file to retrieve
    files only (no subcategories or pages). Handles API continuation
    automatically so all members are returned regardless of category size.

    Args:
        session: Authenticated requests.Session for Commons.
        category: Full category name including 'Category:' prefix.

    Returns:
        list[dict]: Each dict contains 'pageid' (str), 'title' (str,
                    e.g. 'File:Example.jpg'), and 'mid' (str, e.g. 'M12345').
    """
    files = []
    params = {
        "action": "query",
        "generator": "categorymembers",
        "gcmtitle": category,
        "gcmtype": "file",
        "gcmlimit": "max",
        "format": "json",
    }

    while True:
        resp = session.get(COMMONS_API, params=params).json()
        pages = resp.get("query", {}).get("pages", {})

        for pageid, page in pages.items():
            files.append({
                "pageid": pageid,
                "title": page["title"],
                "mid": f"M{pageid}",
            })

        if "continue" in resp:
            params.update(resp["continue"])
            time.sleep(1.5)
        else:
            break

    print(f"  Found {len(files)} files in category")
    return files


def check_for_depicts(session, files, batch_size=25):
    """Extract P180 (depicts) Q-ids and file captions from Commons MediaInfo entities.

    Calls wbgetentities on the Commons API in batches with props=claims|labels.
    Commons MediaInfo entities store structured data under "statements" (not
    "claims"), so the function checks both keys for compatibility. Labels on
    MediaInfo entities are the file captions in each language.

    Includes retry logic (up to 3 attempts per batch) for transient
    HTTP or API errors.

    Args:
        session: Authenticated requests.Session for Commons.
        files: List of file dicts, each containing a 'mid' key (e.g. 'M12345').
        batch_size: Number of M-ids per wbgetentities request (max 50).

    Returns:
        tuple: (depicts_map, captions_map)
            depicts_map: Mapping of M-id to list of Q-ids found in P180.
            captions_map: Mapping of M-id to {'en': str, 'nl': str} captions.
    """
    depicts_map = {}
    captions_map = {}
    mids = [f["mid"] for f in files]

    for i in range(0, len(mids), batch_size):
        batch = mids[i:i + batch_size]
        params = {
            "action": "wbgetentities",
            "ids": "|".join(batch),
            "props": "claims|labels",
            "format": "json",
        }

        # Retry up to 3 times on failure
        resp_json = None
        for attempt in range(3):
            resp = session.get(COMMONS_API, params=params)
            if resp.status_code != 200:
                print(f"  WARNING: Batch {i} attempt {attempt+1}: HTTP {resp.status_code}, retrying...")
                time.sleep(3)
                continue
            resp_json = resp.json()
            if "error" in resp_json:
                print(f"  WARNING: Batch {i} attempt {attempt+1}: API error: {resp_json['error']}, retrying...")
                time.sleep(3)
                continue
            break

        if resp_json is None or "error" in resp_json:
            print(f"  ERROR: Batch {i} failed after 3 attempts, skipping {len(batch)} files")
            continue

        entities = resp_json.get("entities", {})

        # Debug: inspect first entity in first batch
        if i == 0 and entities:
            first_mid = next(iter(entities))
            first_entity = entities[first_mid]
            print(f"  DEBUG: First entity keys: {list(first_entity.keys())}")
            has_statements = "statements" in first_entity
            has_claims = "claims" in first_entity
            print(f"  DEBUG: has 'statements': {has_statements}, has 'claims': {has_claims}")
            if has_statements:
                print(f"  DEBUG: statement properties: {list(first_entity['statements'].keys())}")

        for mid, entity in entities.items():
            # Commons MediaInfo uses "statements", not "claims"
            claims = entity.get("statements", {}) or entity.get("claims", {})

            p180_values = []
            if "P180" in claims:
                for claim in claims["P180"]:
                    mainsnak = claim.get("mainsnak", {})
                    datavalue = mainsnak.get("datavalue", {})
                    if datavalue.get("type") == "wikibase-entityid":
                        qid = datavalue.get("value", {}).get("id", "")
                        if qid:
                            p180_values.append(qid)
            depicts_map[mid] = p180_values

            # Extract EN and NL captions (labels on MediaInfo entities)
            labels = entity.get("labels", {})
            captions_map[mid] = {
                "en": labels.get("en", {}).get("value", ""),
                "nl": labels.get("nl", {}).get("value", ""),
            }

        print(f"  Batch {i}-{i+len(batch)}: checked {len(batch)} files ({len(entities)} returned)")
        time.sleep(1.5)

    with_depicts = sum(1 for v in depicts_map.values() if v)
    print(f"  Files with depicts (P180): {with_depicts}")
    return depicts_map, captions_map


def get_file_details(session, files, batch_size=25):
    """Fetch categories and global usage for files.

    Args:
        session: Authenticated requests session.
        files: List of file dicts with 'title' key.
        batch_size: Number of files per API request.

    Returns:
        dict: Mapping of title to {'categories': [...], 'globalusage': [...]}.
    """
    file_details = {}

    for i in range(0, len(files), batch_size):
        batch = files[i:i + batch_size]
        titles = "|".join(f["title"] for f in batch)
        params = {
            "action": "query",
            "titles": titles,
            "prop": "categories|globalusage",
            "clshow": "!hidden",
            "cllimit": "max",
            "gulimit": "max",
            "format": "json",
        }

        # Handle continuation
        while True:
            resp = session.get(COMMONS_API, params=params).json()
            pages = resp.get("query", {}).get("pages", {})

            for pageid, page in pages.items():
                title = page["title"]
                if title not in file_details:
                    file_details[title] = {"categories": [], "globalusage": []}

                for c in page.get("categories", []):
                    cat_name = c["title"].replace("Category:", "", 1)
                    if cat_name not in file_details[title]["categories"]:
                        file_details[title]["categories"].append(cat_name)

                for u in page.get("globalusage", []):
                    entry = {"wiki": u["wiki"], "title": u["title"]}
                    if entry not in file_details[title]["globalusage"]:
                        file_details[title]["globalusage"].append(entry)

            if "continue" in resp:
                params.update(resp["continue"])
                time.sleep(1.0)
            else:
                break

        print(f"  Batch {i}-{i+len(batch)} done")
        time.sleep(1.5)

    return file_details


def get_depicts_info(qids, batch_size=50):
    """Fetch human filter + EN/NL labels/descriptions for depicts Q-ids in one pass.

    For each Q-id, fetches claims (to check P31=Q5) and labels/descriptions
    in both English and Dutch, all in a single API call per batch.

    Args:
        qids: List of Q-ids (e.g., ['Q123', 'Q456']).
        batch_size: Number of items per API request.

    Returns:
        tuple: (human_qids: set, info_en: dict, info_nl: dict)
            human_qids: Set of Q-ids that are instances of Q5 (human).
            info_en: Mapping of Q-id to {'label': str, 'description': str} in English.
            info_nl: Mapping of Q-id to {'label': str, 'description': str} in Dutch.
    """
    unique_qids = list(set(qids))
    human_qids = set()
    info_en = {}
    info_nl = {}

    if not unique_qids:
        return human_qids, info_en, info_nl

    print(f"  Fetching claims + EN/NL labels for {len(unique_qids)} depicts Q-ids...")
    headers = {"User-Agent": USER_AGENT}

    for i in range(0, len(unique_qids), batch_size):
        batch = unique_qids[i:i + batch_size]
        params = {
            "action": "wbgetentities",
            "ids": "|".join(batch),
            "props": "claims|labels|descriptions",
            "languages": "en|nl",
            "format": "json",
        }

        resp = requests.get(WIKIDATA_API, params=params, headers=headers)
        if resp.status_code != 200:
            print(f"  Wikidata API error: {resp.status_code}")
            continue

        data = resp.json()
        entities = data.get("entities", {})

        for qid, entity in entities.items():
            if "missing" in entity:
                continue

            # Check P31 (instance of) for Q5 (human)
            claims = entity.get("claims", {})
            for claim in claims.get("P31", []):
                mainsnak = claim.get("mainsnak", {})
                datavalue = mainsnak.get("datavalue", {})
                if datavalue.get("type") == "wikibase-entityid":
                    if datavalue.get("value", {}).get("id", "") == "Q5":
                        human_qids.add(qid)
                        break

            # Extract EN and NL labels/descriptions
            labels = entity.get("labels", {})
            descs = entity.get("descriptions", {})
            info_en[qid] = {
                "label": labels.get("en", {}).get("value", ""),
                "description": descs.get("en", {}).get("value", ""),
            }
            info_nl[qid] = {
                "label": labels.get("nl", {}).get("value", ""),
                "description": descs.get("nl", {}).get("value", ""),
            }

        print(f"  Batch {i}-{i+len(batch)}: {len(batch)} checked")
        time.sleep(0.5)

    print(f"  Found {len(human_qids)} humans out of {len(unique_qids)} Q-ids")
    return human_qids, info_en, info_nl


def get_wikidata_labels(qids, batch_size=50, language="en"):
    """Fetch labels and descriptions for Wikidata items in a single language.

    Used for the wikidata_usage Q-ids (global-usage links). For depicts
    Q-ids, prefer get_depicts_info() which fetches both languages and
    the human filter in one pass.

    Deduplicates input Q-ids before querying.

    Args:
        qids: List of Q-ids (e.g., ['Q123', 'Q456']).
        batch_size: Number of items per wbgetentities request (max 50).
        language: BCP-47 language code (default 'en').

    Returns:
        dict: Mapping of Q-id to {'label': str, 'description': str}.
    """
    qid_info = {}
    unique_qids = list(set(qids))

    if not unique_qids:
        print("  No Q-ids to fetch labels for")
        return qid_info

    print(f"  Fetching {language} labels for {len(unique_qids)} Wikidata items: {unique_qids[:5]}...")

    headers = {"User-Agent": USER_AGENT}

    for i in range(0, len(unique_qids), batch_size):
        batch = unique_qids[i:i + batch_size]
        params = {
            "action": "wbgetentities",
            "ids": "|".join(batch),
            "props": "labels|descriptions",
            "languages": language,
            "format": "json",
        }

        resp = requests.get(WIKIDATA_API, params=params, headers=headers)
        if resp.status_code != 200:
            print(f"  Wikidata API error: {resp.status_code}")
            continue

        data = resp.json()
        entities = data.get("entities", {})

        for qid, entity in entities.items():
            if "missing" in entity:
                continue
            label = entity.get("labels", {}).get(language, {}).get("value", "")
            description = entity.get("descriptions", {}).get(language, {}).get("value", "")
            qid_info[qid] = {"label": label, "description": description}

        print(f"  Batch {i}-{i+len(batch)}: got {len(entities)} entities")
        time.sleep(0.5)

    return qid_info


def main():
    """Run the full pipeline: fetch, filter, and export portrait data.

    Steps:
        0. Authenticate with Wikimedia Commons.
        1. Fetch all files from the "Portraits from Koninklijke Bibliotheek" category.
        2. Extract depicts (P180) Q-ids and EN/NL file captions from each
           file's MediaInfo (single API call per batch).
        2b. Query Wikidata for all depicts Q-ids in one pass: determine which
            are humans (P31=Q5) and retrieve EN + NL labels/descriptions.
            Discard non-human depicts items.
        3. Split files into with/without depicts.
        4. Fetch non-hidden categories and global usage for all files.
        5. Fetch EN labels for Wikidata-usage Q-ids.
        6. Build result rows — one row per depicted human per file, including
           EN/NL captions.
        7. Write CSV files (kb-portraits-depicts-q5-status-all-YYYYMMDD.csv,
           kb-portraits-depicts-q5-status-missing-YYYYMMDD.csv).
        8. Write Excel workbook (kb-portraits-depicts-q5-status-YYYYMMDD.xlsx)
           with three sheets: all, missing, present.
    """

    # Step 0: Authenticate
    print("=== Step 0: Authenticating ===")
    if not USERNAME or not PASSWORD:
        print("Error: COMMONS_USERNAME and COMMONS_PASSWORD must be set in .env file")
        return

    session = create_authenticated_session()
    if not session:
        print("Authentication failed. Exiting.")
        return

    # Step 1: Get all files from category
    print("\n=== Step 1: Fetching category members ===")
    files = get_category_members(session)

    if not files:
        print("No files found in category.")
        return

    # Step 2: Check which files have depicts (P180)
    print("\n=== Step 2: Checking for P180 statements ===")
    depicts_map, captions_map = check_for_depicts(session, files)

    # Debug: verify depicts_map
    print(f"\n  DEBUG: depicts_map has {len(depicts_map)} entries")
    files_with_p180 = [mid for mid, vals in depicts_map.items() if vals]
    print(f"  DEBUG: {len(files_with_p180)} files have P180 values")
    if files_with_p180:
        print(f"  DEBUG: Example with P180: {files_with_p180[0]} -> {depicts_map[files_with_p180[0]]}")

    # Step 2b: Filter depicts to humans (Q5) + fetch EN/NL labels in one pass
    print("\n=== Step 2b: Filtering depicts to humans + fetching labels ===")
    all_depicts_qids = []
    for vals in depicts_map.values():
        all_depicts_qids.extend(vals)
    human_qids, depicts_qid_info_en, depicts_qid_info_nl = get_depicts_info(all_depicts_qids)

    # Remove non-human Q-ids from depicts_map
    for mid in depicts_map:
        depicts_map[mid] = [qid for qid in depicts_map[mid] if qid in human_qids]

    files_with_p180_human = [mid for mid, vals in depicts_map.items() if vals]
    print(f"  Files with depicts (P180, humans only): {len(files_with_p180_human)}")

    # Step 3: Split into files with and without depicts
    no_depicts_files = [f for f in files if not depicts_map.get(f["mid"], [])]
    with_depicts_files = [f for f in files if depicts_map.get(f["mid"], [])]
    print(f"\n  Files WITH depicts (P180): {len(with_depicts_files)}")
    print(f"  Files WITHOUT depicts (P180): {len(no_depicts_files)}")

    # Step 4: Get categories and global usage for ALL files
    print("\n=== Step 4: Fetching categories and global usage ===")
    details = get_file_details(session, files)

    # Step 5: Collect all Wikidata Q-ids and fetch labels/descriptions
    print("\n=== Step 5: Fetching Wikidata labels ===")
    all_qids = []
    for f in files:
        info = details.get(f["title"], {})
        for u in info.get("globalusage", []):
            if u["wiki"] == "www.wikidata.org" and u["title"].startswith("Q") and u["title"][1:].isdigit():
                all_qids.append(u["title"])

    qid_info = get_wikidata_labels(all_qids)

    # Step 6: Build results — one row per depicts Q-id (split multi-value rows)
    print("\n=== Step 6: Building results (one row per depicts item) ===")
    all_results = []
    for f in files:
        info = details.get(f["title"], {})
        # Filter to only Wikidata Q-ids (not project pages etc.)
        wikidata_usage = [
            u["title"] for u in info.get("globalusage", [])
            if u["wiki"] == "www.wikidata.org" and u["title"].startswith("Q") and u["title"][1:].isdigit()
        ]
        labels = [qid_info.get(qid, {}).get("label", "") for qid in wikidata_usage]

        # Shared fields across split rows
        captions = captions_map.get(f["mid"], {})
        shared = {
            "file": f"https://commons.wikimedia.org/entity/{f['mid']}",
            "title": f["title"],
            "file_caption_en": captions.get("en", ""),
            "file_caption_nl": captions.get("nl", ""),
            "categories": "; ".join(info.get("categories", [])),
            "wikidata_usage": "; ".join(wikidata_usage),
            "wikidata_label": "; ".join(labels),
        }

        depicts_values = depicts_map.get(f["mid"], [])

        if depicts_values:
            # One row per depicted human
            for qid in depicts_values:
                row = dict(shared)
                row["depicts_P180"] = qid
                row["depicts_P180_label_en"] = depicts_qid_info_en.get(qid, {}).get("label", "")
                row["depicts_P180_description_en"] = depicts_qid_info_en.get(qid, {}).get("description", "")
                row["depicts_P180_label_nl"] = depicts_qid_info_nl.get(qid, {}).get("label", "")
                row["depicts_P180_description_nl"] = depicts_qid_info_nl.get(qid, {}).get("description", "")
                all_results.append(row)
        else:
            # No depicts — single row with empty depicts fields
            row = dict(shared)
            row["depicts_P180"] = ""
            row["depicts_P180_label_en"] = ""
            row["depicts_P180_description_en"] = ""
            row["depicts_P180_label_nl"] = ""
            row["depicts_P180_description_nl"] = ""
            all_results.append(row)

    # Split results
    results_no_depicts = [r for r in all_results if not r["depicts_P180"]]
    results_with_depicts = [r for r in all_results if r["depicts_P180"]]

    print(f"  Total rows (after split): {len(all_results)}")
    print(f"  Rows with P180: {len(results_with_depicts)}")
    print(f"  Rows without P180: {len(results_no_depicts)}")

    # Step 7: Write CSV files
    print("\n=== Step 7: Writing CSV files ===")
    datestamp = date.today().strftime("%Y%m%d")
    fieldnames = ["file", "title", "file_caption_en", "file_caption_nl",
                   "depicts_P180", "depicts_P180_label_en", "depicts_P180_description_en",
                   "depicts_P180_label_nl", "depicts_P180_description_nl", "categories",
                   "wikidata_usage", "wikidata_label"]

    csv_all = f"kb-portraits-depicts-q5-status-all-{datestamp}.csv"
    with open(csv_all, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_results)
    print(f"  Saved {len(all_results)} rows to {csv_all}")

    csv_missing = f"kb-portraits-depicts-q5-status-missing-{datestamp}.csv"
    with open(csv_missing, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results_no_depicts)
    print(f"  Saved {len(results_no_depicts)} rows to {csv_missing}")

    # Step 8: Write Excel file
    print("\n=== Step 8: Writing Excel file ===")

    def write_excel_sheet(wb, sheet_name, rows):
        ws = wb.create_sheet(title=sheet_name)
        # Write header
        for col_idx, name in enumerate(fieldnames, 1):
            ws.cell(row=1, column=col_idx, value=name)

        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, name in enumerate(fieldnames, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(name, ""))

    wb = Workbook()
    wb.remove(wb.active)  # Remove default empty sheet
    write_excel_sheet(wb, "all", all_results)
    write_excel_sheet(wb, "missing", results_no_depicts)
    write_excel_sheet(wb, "present", results_with_depicts)
    xlsx_file = f"kb-portraits-depicts-q5-status-{datestamp}.xlsx"
    wb.save(xlsx_file)
    print(f"  Saved {xlsx_file} (3 sheets: all / missing / present)")

    print("\nDone!")


if __name__ == "__main__":
    main()
